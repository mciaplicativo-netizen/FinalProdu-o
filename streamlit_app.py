"""
Production & Inventory - Dashboard v3
Changes requested by user:
- Sidebar navigation (Produção / Estoque MP / Estoque Injetados / Apontamentos Online)
- Produção page: efficiency by Operador (%), product summary (Programado/Realizado/Perda),
  filters including Funcionário (Operador), totals Kg, Kg Aparas, Observações, Tabela Produção editable.
- Estoque Injetados: totalizer per product, movement input (entrada/saída) and history
- Apontamentos Online: per-machine quick status with colored boxes
- SQLite <-> Excel synchronization preserved
"""

import streamlit as st
import pandas as pd
import sqlite3
from pathlib import Path
import time, datetime, json
from io import BytesIO
from openpyxl import load_workbook

APP_DIR = Path(".")
DB_PATH = APP_DIR / "database.db"
XLSX_PATH = APP_DIR / "Indicadores_CPP1.xlsx"
LOCK_PATH = APP_DIR / ".write_lock"

st.set_page_config(page_title="Produção & Estoque - Dashboard v4", layout="wide")

st.markdown(
    '''
    <style>
    body {font-family: 'Inter', Arial, sans-serif;}
    div[data-testid="stMetricValue"] {font-size: 26px !important;}
    .metric-card {background: #f7fafc; border-radius: 12px; padding: 10px; text-align: center; box-shadow: 0 2px 5px rgba(0,0,0,0.05);}
    .status-dot {width:14px;height:14px;border-radius:50%;display:inline-block;margin-right:6px;}
    </style>
    ''',
    unsafe_allow_html=True
)


# ------------------ Helpers ------------------
def with_lock(fn):
    def wrapper(*args, **kwargs):
        while LOCK_PATH.exists():
            time.sleep(0.05)
        try:
            LOCK_PATH.write_text("lock")
            return fn(*args, **kwargs)
        finally:
            if LOCK_PATH.exists():
                try:
                    LOCK_PATH.unlink()
                except:
                    pass
    return wrapper

@with_lock
def write_excel_sheets(sheet_dfs: dict):
    # writes/overwrites the given sheets into the Excel file, preserving others
    if not XLSX_PATH.exists():
        with pd.ExcelWriter(XLSX_PATH, engine="openpyxl") as writer:
            for name, df in sheet_dfs.items():
                df.to_excel(writer, sheet_name=name, index=False)
        return
    wb = load_workbook(XLSX_PATH)
    # remove sheets to overwrite
    for name in list(sheet_dfs.keys()):
        if name in wb.sheetnames:
            std = wb[name]
            wb.remove(std)
    wb.save(XLSX_PATH)
    with pd.ExcelWriter(XLSX_PATH, engine="openpyxl", mode="a") as writer:
        for name, df in sheet_dfs.items():
            df.to_excel(writer, sheet_name=name, index=False)

def read_table(table):
    conn = sqlite3.connect(DB_PATH)
    try:
        df = pd.read_sql_query(f"SELECT * FROM [{table}]", conn)
    except Exception:
        df = pd.DataFrame()
    conn.close()
    return df

def write_table(table, df):
    conn = sqlite3.connect(DB_PATH)
    df.to_sql(table, conn, if_exists="replace", index=False)
    conn.close()

def init_db_from_excel(mapping):
    conn = sqlite3.connect(DB_PATH)
    for sheet, table in mapping.items():
        try:
            df = pd.read_excel(XLSX_PATH, sheet_name=sheet)
        except Exception:
            continue
        df.to_sql(table, conn, if_exists="replace", index=False)
    conn.close()

# ------------------ Detect sheets ------------------
detected = []
if XLSX_PATH.exists():
    try:
        x = pd.ExcelFile(XLSX_PATH)
        detected = x.sheet_names
    except Exception:
        detected = []

default_mapping = {
    "Estoque MP": "estoque_mp",
    "Estoque Injetados": "estoque_injetados",
    "Produção - injeção+ Zamac": "producao"
}

st.sidebar.title("Navegação")
menu = st.sidebar.radio("Ir para", ["Produção", "Estoque MP", "Estoque Injetados", "Apontamentos Online"])

st.sidebar.markdown("---")
st.sidebar.header("Mapeamento de sheets (ajuste se necessário)")
user_mapping = {}
for sx, tbl in default_mapping.items():
    val = sx if sx in detected else sx
    newname = st.sidebar.text_input(f"Sheet fonte para '{tbl}'", value=val, key="map_"+tbl)
    user_mapping[newname] = tbl

st.sidebar.markdown("---")
st.sidebar.header("Import / Export")
uploaded = st.sidebar.file_uploader("Substituir arquivo Excel (opcional)", type=["xlsx","xls"])
if uploaded is not None:
    with open(XLSX_PATH, "wb") as f:
        f.write(uploaded.getbuffer())
    st.sidebar.success("Arquivo Excel substituído. Recarregue para aplicar.")

if st.sidebar.button("Sincronizar Excel → DB (forçar)"):
    init_db_from_excel({k:v for k,v in user_mapping.items()})
    st.sidebar.success("Sincronização inicial (Excel→DB) feita. Recarregue a página.")

# Ensure DB exists (import if necessary)
if not DB_PATH.exists() and XLSX_PATH.exists():
    init_db_from_excel({k:v for k,v in user_mapping.items()})

# Load tables
tables = {}
for sheet_name, table_name in user_mapping.items():
    tables[table_name] = read_table(table_name)

# Coerce production date if present
prod_df = tables.get("producao", pd.DataFrame())
if "Data" in prod_df.columns:
    try:
        prod_df["Data"] = pd.to_datetime(prod_df["Data"], errors="coerce").dt.date
    except Exception:
        pass

# ------------------ Produção ------------------
if menu == "Produção":
    st.title("Produção — Dashboard")
    st.markdown("KPIs e visão rápida da produção.")

    left, right = st.columns([1,3], gap="large")
    with left:
        st.subheader("Filtros")
        # date range
        if "Data" in prod_df.columns:
            min_date = pd.to_datetime(prod_df["Data"]).min().date() if not prod_df["Data"].isna().all() else None
            max_date = pd.to_datetime(prod_df["Data"]).max().date() if not prod_df["Data"].isna().all() else None
            date_range = st.date_input("Período", value=(min_date, max_date) if min_date and max_date else None)
        else:
            date_range = None
        machines = sorted(prod_df["Máquina"].dropna().unique().tolist()) if "Máquina" in prod_df.columns else []
        sel_machine = st.multiselect("Máquina", options=machines, default=machines)
        products = sorted(prod_df["Produto"].dropna().unique().tolist()) if "Produto" in prod_df.columns else []
        sel_product = st.multiselect("Produto", options=products, default=products)
        turns = sorted(prod_df["Turno"].dropna().unique().tolist()) if "Turno" in prod_df.columns else []
        sel_turn = st.multiselect("Turno", options=turns, default=turns)
        # Funcionário filter
        funcionarios = sorted(prod_df["Operador"].dropna().unique().tolist()) if "Operador" in prod_df.columns else []
        sel_func = st.multiselect("Funcionário (Operador)", options=funcionarios, default=funcionarios)

    with right:
        df = prod_df.copy()
        if df.empty:
            st.info("Nenhum dado de produção disponível. Importe do Excel no sidebar ou sincronize.")
        else:
            # apply filters
            if date_range and isinstance(date_range, tuple) and len(date_range)==2:
                start, end = date_range
                df = df[(pd.to_datetime(df["Data"]) >= pd.to_datetime(start)) & (pd.to_datetime(df["Data"]) <= pd.to_datetime(end))]
            if "Máquina" in df.columns and sel_machine:
                df = df[df["Máquina"].isin(sel_machine)]
            if "Produto" in df.columns and sel_product:
                df = df[df["Produto"].isin(sel_product)]
            if "Turno" in df.columns and sel_turn:
                df = df[df["Turno"].isin(sel_turn)]
            if "Operador" in df.columns and sel_func:
                df = df[df["Operador"].isin(sel_func)]

            # KPIs
            col1, col2, col3, col4 = st.columns(4)
            total_prod = int(df["Realizado"].sum()) if "Realizado" in df.columns else 0
            eficiencia_mean = float(df["Eficiência"].mean()) if "Eficiência" in df.columns and not df["Eficiência"].isna().all() else None
            total_ciclos = int(df["Ciclos"].sum()) if "Ciclos" in df.columns else None
            aparas = float(df["Kg Aparas"].sum()) if "Kg Aparas" in df.columns else None
            
# KPIs visuais (cards)
total_prod = int(df["Realizado"].sum()) if "Realizado" in df.columns else 0
eficiencia_mean = float(df["Eficiência"].mean()) if "Eficiência" in df.columns and not df["Eficiência"].isna().all() else None
total_ciclos = int(df["Ciclos"].sum()) if "Ciclos" in df.columns else None
aparas = float(df["Kg Aparas"].sum()) if "Kg Aparas" in df.columns else None

k1, k2, k3, k4 = st.columns(4)
k1.markdown(f"<div class='metric-card'><div>Total Produzido</div><div style='font-size:22px;font-weight:700'>{total_prod:,d}</div></div>", unsafe_allow_html=True)
k2.markdown(f"<div class='metric-card'><div>Eficiência Média</div><div style='font-size:22px;font-weight:700'>{(eficiencia_mean*100):.2f}%</div></div>", unsafe_allow_html=True) if eficiencia_mean else k2.markdown("<div class='metric-card'>—</div>", unsafe_allow_html=True)
k3.markdown(f"<div class='metric-card'><div>Total de Ciclos</div><div style='font-size:22px;font-weight:700'>{total_ciclos if total_ciclos else '—'}</div></div>", unsafe_allow_html=True)
k4.markdown(f"<div class='metric-card'><div>Kg Aparas</div><div style='font-size:22px;font-weight:700'>{aparas:.2f if aparas else '—'}</div></div>", unsafe_allow_html=True)
            # Totais de Kg
total_kg_pecas = float(df["Kg"].sum()) if "Kg" in df.columns else None
            st.write("**Totais (Kg)**")
            st.write(f"Kg de Peças: {total_kg_pecas:.2f}" if total_kg_pecas is not None else "Kg de Peças: —")

            # Observações
            st.subheader("Apontamentos / Observações")
            if "Observações" in df.columns:
                obs = df[~df["Observações"].isna()][["Data","Máquina","Produto","Operador","Observações"]].copy()
                if not obs.empty:
                    st.dataframe(obs.sort_values("Data", ascending=False).reset_index(drop=True))
                else:
                    st.info("Nenhuma observação registrada no período.")
            else:
                st.info("Coluna 'Observações' não encontrada.")

            # Efficiency by operator chart (percent)
            st.subheader("Eficiência por Operador (%)")
            if "Operador" in df.columns and "Eficiência" in df.columns:
                eff = df.groupby("Operador", as_index=False)["Eficiência"].mean().sort_values("Eficiência", ascending=False)
                eff["Eficiência"] = eff["Eficiência"] * 100
                st.bar_chart(eff.set_index("Operador")["Eficiência"])
            else:
                st.info("Colunas 'Operador' e/ou 'Eficiência' não encontradas para gerar gráfico.")

            # Product summary: Produto x Programado x Realizado x Perda
            st.subheader("Resumo por Produto (Programado / Realizado / Perda)")
            if "Produto" in df.columns and "Programado" in df.columns and "Realizado" in df.columns:
                prod_sum = df.groupby("Produto", as_index=False).agg({"Programado":"sum","Realizado":"sum"})
                prod_sum["Perda"] = (prod_sum["Programado"] - prod_sum["Realizado"]).clip(lower=0)
                prod_sum = prod_sum.sort_values("Programado", ascending=False)
                st.table(prod_sum.rename(columns={"Programado":"Programado","Realizado":"Realizado","Perda":"Perda"}))
            else:
                st.info("Colunas necessárias para resumo por produto não encontradas.")

            # Tabela Produção (editable)
            st.subheader("Tabela Produção")
            edited = st.data_editor(df, num_rows="dynamic", use_container_width=True)

            csv_data = df.to_csv(index=False).encode("utf-8")
            st.download_button("Exportar Tabela Produção (CSV)", data=csv_data, file_name="tabela_producao_filtrada.csv", mime="text/csv")
            if st.button("Salvar alterações na produção"):
                write_table("producao", edited)
                # write back to excel sheet (find sheet name mapping)
                try:
                    sheet_name = list(user_mapping.keys())[list(user_mapping.values()).index("producao")]
                except Exception:
                    sheet_name = "Produção - injeção+ Zamac"
                write_excel_sheets({sheet_name: edited})
                st.success("Produção salva no DB e Excel.")

# ------------------ Estoque MP ------------------
elif menu == "Estoque MP":
    st.title("Estoque MP")
    df_mp = tables.get("estoque_mp", pd.DataFrame(columns=["mp_id","mp_nome","quantidade","unidade","local"]))
    st.write("Edite o estoque MP abaixo:")
    edited_mp = st.data_editor(df_mp, num_rows="dynamic", use_container_width=True)
    if st.button("Salvar Estoque MP"):
        write_table("estoque_mp", edited_mp)
        try:
            sheet_name = list(user_mapping.keys())[list(user_mapping.values()).index("estoque_mp")]
        except Exception:
            sheet_name = "Estoque MP"
        write_excel_sheets({sheet_name: edited_mp})
        st.success("Estoque MP salvo no DB e Excel.")

# ------------------ Estoque Injetados ------------------
elif menu == "Estoque Injetados":
    st.title("Estoque Injetados")
    # Ensure movimentation table exists
    conn = sqlite3.connect(DB_PATH)
    conn.execute('CREATE TABLE IF NOT EXISTS movimentacao_injetados (id INTEGER PRIMARY KEY AUTOINCREMENT, sku TEXT, nome TEXT, qty REAL, motivo TEXT, operador TEXT, data TEXT)')
    conn.commit()
    conn.close()

    # Totals per product (top area)
    df_inj = tables.get("estoque_injetados", pd.DataFrame(columns=["sku","nome","quantidade","unidade","local"]))
    if not df_inj.empty and "sku" in df_inj.columns:
        totals = df_inj.groupby(["sku","nome"], as_index=False)["quantidade"].sum()
    else:
        totals = pd.DataFrame(columns=["sku","nome","quantidade"])
    st.subheader("Totalizador de Estoque por Produto")
    st.table(totals.sort_values("quantidade", ascending=False))

    st.subheader("Movimentação (lançar entradas/saídas)")
    with st.form("mov_inj_form"):
        sku = st.text_input("SKU / Código")
        nome = st.text_input("Nome do produto")
        qty = st.number_input("Quantidade (positivo = entrada, negativo = saída)", value=0.0)
        motivo = st.text_input("Motivo")
        operador = st.text_input("Operador")
        submit_mov = st.form_submit_button("Registrar movimentação")
        if submit_mov:
            conn = sqlite3.connect(DB_PATH)
            conn.execute('INSERT INTO movimentacao_injetados (sku,nome,qty,motivo,operador,data) VALUES (?,?,?,?,?,?)', (sku,nome,float(qty),motivo,operador,datetime.datetime.now().isoformat()))
            conn.commit()
            conn.close()
            # update totals in estoque_injetados table accordingly
            df = read_table('estoque_injetados')
            if df.empty:
                df = pd.DataFrame(columns=["sku","nome","quantidade","unidade","local"])
            if 'sku' in df.columns and sku:
                mask = df['sku'].astype(str) == str(sku)
                if mask.any():
                    df.loc[mask, 'quantidade'] = df.loc[mask, 'quantidade'].astype(float) + float(qty)
                else:
                    newrow = {"sku":sku, "nome":nome, "quantidade":qty, "unidade":"", "local":""}
                    df = pd.concat([df, pd.DataFrame([newrow])], ignore_index=True)
            else:
                newrow = {"sku":sku, "nome":nome, "quantidade":qty, "unidade":"", "local":""}
                df = pd.concat([df, pd.DataFrame([newrow])], ignore_index=True)
            write_table('estoque_injetados', df)
            try:
                sheet_name = list(user_mapping.keys())[list(user_mapping.values()).index("estoque_injetados")]
            except Exception:
                sheet_name = "Estoque Injetados"
            write_excel_sheets({sheet_name: df})
            st.success('Movimentação registrada e estoque atualizado.')

    st.subheader('Histórico de Movimentações')
    conn = sqlite3.connect(DB_PATH)
    mov = pd.read_sql_query('SELECT * FROM movimentacao_injetados ORDER BY data DESC', conn)
    conn.close()
    st.dataframe(mov)

# ------------------ Apontamentos Online ------------------
elif menu == "Apontamentos Online":
    st.title("Apontamentos Online")
    st.markdown("Atualize rapidamente o que está rodando por máquina. Status muda de cor visualmente.")
    machines = ["Oriente 45", "Oriente 35", "Himaco 80", "Himaco 40", "Jasot", "MG", "Máq. 1 (Zamac)", "Máq. 2 (Zamac)"]

    # ensure table exists
    conn = sqlite3.connect(DB_PATH)
    conn.execute('CREATE TABLE IF NOT EXISTS apontamentos (machine TEXT PRIMARY KEY, produto TEXT, operador TEXT, status TEXT, updated_at TEXT)')
    conn.commit()
    conn.close()

    cols = st.columns(2)
    for i, m in enumerate(machines):
        with cols[i%2]:
            st.subheader(m)
            # load current
            conn = sqlite3.connect(DB_PATH)
            cur = conn.cursor()
            cur.execute('SELECT produto, operador, status, updated_at FROM apontamentos WHERE machine=?', (m,))
            row = cur.fetchone()
            conn.close()
            prod = row[0] if row else ""
            oper = row[1] if row else ""
            status = row[2] if row else "Em Injeção"
            updated = row[3] if row else ""
            # color box based on status
            color = "#28a745" if status=="Em Injeção" else ("#dc3545" if status=="Quebra" else ("#ffc107" if status=="Setup" else "#6c757d"))
            st.markdown(f"<div><span class='status-dot' style='background:{color}'></span><b>{status}</b></div>", unsafe_allow_html=True)
            st.write(f"**Produto:** {prod}")
            st.write(f"**Operador:** {oper}")
            with st.form(f"form_{m}"):
                p = st.text_input("Produto (código)", value=prod, key=f"prod_{m}")
                o = st.text_input("Operador", value=oper, key=f"oper_{m}")
                s = st.selectbox("Status", options=["Em Injeção", "Quebra", "Setup", "Parada"], index=0 if status=="Em Injeção" else 1, key=f"stat_{m}")
                submit = st.form_submit_button("Atualizar")
                if submit:
                    conn = sqlite3.connect(DB_PATH)
                    conn.execute('REPLACE INTO apontamentos (machine, produto, operador, status, updated_at) VALUES (?,?,?,?,?)', (m, p, o, s, datetime.datetime.now().isoformat()))
                    conn.commit()
                    conn.close()
                    st.success("Apontamento atualizado.")

st.sidebar.caption("Use com cuidado: gravação reescreve sheets no Excel. Faça backup antes de usar.")